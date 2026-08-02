"""
Report Service — generates XLSX and JSON reports.
"""
from __future__ import annotations
from typing import Dict, Any
from datetime import datetime
from app.core.logging import get_logger

logger = get_logger(__name__)


def generate_xlsx_report(analysis_data: Dict[str, Any]) -> bytes:
    """
    Generate a professional Excel report. Returns bytes.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Color palette
        HEADER_COLOR = "064DEB"
        ACCENT_COLOR = "FFC72C"
        LIGHT_BG = "F5F7FF"
        GREEN = "08AD61"
        RED = "E53935"
        DARK = "071642"

        def style_header(cell, bg=HEADER_COLOR, fg="FFFFFF", bold=True, size=11):
            cell.font = Font(bold=bold, color=fg, size=size)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def style_data(cell, bold=False):
            cell.font = Font(bold=bold, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        def add_section_header(ws, row, title, cols=2):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
            cell = ws.cell(row=row, column=1, value=title)
            style_header(cell, bg=DARK, size=11)
            return row + 1

        def add_kv(ws, row, key, value):
            ws.cell(row=row, column=1, value=key).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=str(value)).font = Font(size=10)
            ws.row_dimensions[row].height = 16
            return row + 1

        # ── Sheet 1: Executive Summary ─────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 55

        scores = analysis_data.get("scores", {})
        summary = analysis_data.get("summary", {})
        video = analysis_data.get("video", {})

        # Title
        ws1.merge_cells("A1:B1")
        t = ws1["A1"]
        t.value = "🎬 Hook AI — Video Analysis Report"
        t.font = Font(bold=True, size=16, color="064DEB")
        t.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 30

        ws1.merge_cells("A2:B2")
        t2 = ws1["A2"]
        t2.value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        t2.font = Font(size=10, italic=True, color="53648B")
        t2.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[2].height = 18

        row = 4
        row = add_section_header(ws1, row, "Video Information", 2)
        row = add_kv(ws1, row, "Title", video.get("title", "N/A"))
        row = add_kv(ws1, row, "Duration", f"{video.get('duration', 0):.1f}s")
        row = add_kv(ws1, row, "Source", video.get("source", "N/A"))
        row = add_kv(ws1, row, "Resolution", f"{video.get('width', 0)}x{video.get('height', 0)}")

        row += 1
        row = add_section_header(ws1, row, "Overall Scores", 2)
        row = add_kv(ws1, row, "Overall Score", f"{summary.get('overall_score', 0):.1f} / 10")
        row = add_kv(ws1, row, "Rating", summary.get("rating", "N/A"))
        row = add_kv(ws1, row, "Summary", summary.get("summary", ""))

        row += 1
        row = add_section_header(ws1, row, "Score Breakdown", 2)
        for key, label in [("hook","Hook"), ("cta","CTA"), ("tone","Tone"), ("visual","Visual"),
                            ("pacing","Pacing"), ("clarity","Clarity"), ("engagement","Engagement")]:
            row = add_kv(ws1, row, label, f"{scores.get(key, 0):.1f} / 10")

        # ── Sheet 2: Hooks ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Hooks")
        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 50
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 10
        ws2.column_dimensions["E"].width = 12
        ws2.column_dimensions["F"].width = 12

        for col, header in enumerate(["#", "Hook Text", "Start", "End", "Type", "Score"], 1):
            c = ws2.cell(row=1, column=col, value=header)
            style_header(c)

        hooks = analysis_data.get("hooks", {}).get("hooks", [])
        for i, hook in enumerate(hooks[:20], 1):
            ws2.cell(row=i+1, column=1, value=i)
            ws2.cell(row=i+1, column=2, value=hook.get("text", ""))
            ws2.cell(row=i+1, column=3, value=f"{hook.get('start', 0):.1f}s")
            ws2.cell(row=i+1, column=4, value=f"{hook.get('end', 0):.1f}s")
            ws2.cell(row=i+1, column=5, value=hook.get("type", ""))
            ws2.cell(row=i+1, column=6, value=f"{hook.get('score', 0):.0f}/100")

        # ── Sheet 3: CTAs ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("CTAs")
        ws3.column_dimensions["A"].width = 10
        ws3.column_dimensions["B"].width = 50
        for col, header in enumerate(["#", "CTA Text", "Start", "End", "Type", "Strength", "Score"], 1):
            c = ws3.cell(row=1, column=col, value=header)
            style_header(c)

        ctas = analysis_data.get("cta", {}).get("ctas", [])
        for i, cta in enumerate(ctas[:20], 1):
            ws3.cell(row=i+1, column=1, value=i)
            ws3.cell(row=i+1, column=2, value=cta.get("text", ""))
            ws3.cell(row=i+1, column=3, value=f"{cta.get('start', 0):.1f}s")
            ws3.cell(row=i+1, column=4, value=f"{cta.get('end', 0):.1f}s")
            ws3.cell(row=i+1, column=5, value=cta.get("type", ""))
            ws3.cell(row=i+1, column=6, value=cta.get("strength", ""))
            ws3.cell(row=i+1, column=7, value=f"{cta.get('score', 0):.0f}/100")

        # ── Sheet 4: Tone & Emotion ────────────────────────────────────────
        ws4 = wb.create_sheet("Tone & Emotion")
        ws4.column_dimensions["A"].width = 28
        ws4.column_dimensions["B"].width = 30

        tone_data = analysis_data.get("tone", {})
        row4 = 1
        for key, val in [
            ("Primary Tone", tone_data.get("primary_tone", "")),
            ("Sentiment", tone_data.get("sentiment", "")),
            ("Energy Score", f"{tone_data.get('energy_score', 0):.1f}/100"),
            ("Clarity Score", f"{tone_data.get('clarity_score', 0):.1f}/100"),
            ("Confidence Score", f"{tone_data.get('confidence_score', 0):.1f}/100"),
        ]:
            ws4.cell(row=row4, column=1, value=key).font = Font(bold=True)
            ws4.cell(row=row4, column=2, value=val)
            row4 += 1

        row4 += 1
        ws4.cell(row=row4, column=1, value="Emotion Scores").font = Font(bold=True)
        row4 += 1
        for emo, score in (tone_data.get("emotions") or {}).items():
            ws4.cell(row=row4, column=1, value=emo.title())
            ws4.cell(row=row4, column=2, value=f"{score * 100:.1f}%")
            row4 += 1

        # ── Sheet 5: Recommendations ───────────────────────────────────────
        ws5 = wb.create_sheet("Recommendations")
        for col, header in enumerate(["Priority", "Category", "Title", "Description", "Action", "Timestamp"], 1):
            c = ws5.cell(row=1, column=col, value=header)
            style_header(c)
        ws5.column_dimensions["A"].width = 10
        ws5.column_dimensions["B"].width = 12
        ws5.column_dimensions["C"].width = 35
        ws5.column_dimensions["D"].width = 50
        ws5.column_dimensions["E"].width = 55
        ws5.column_dimensions["F"].width = 12

        for i, rec in enumerate(analysis_data.get("recommendations", [])[:20], 2):
            ws5.cell(row=i, column=1, value=rec.get("priority", ""))
            ws5.cell(row=i, column=2, value=rec.get("category", ""))
            ws5.cell(row=i, column=3, value=rec.get("title", ""))
            ws5.cell(row=i, column=4, value=rec.get("description", ""))
            ws5.cell(row=i, column=5, value=rec.get("suggested_action", ""))
            ts = rec.get("timestamp")
            ws5.cell(row=i, column=6, value=f"{ts:.1f}s" if ts is not None else "")

        # ── Sheet 6: Transcript ────────────────────────────────────────────
        ws6 = wb.create_sheet("Transcript")
        for col, header in enumerate(["Segment", "Start", "End", "Text"], 1):
            c = ws6.cell(row=1, column=col, value=header)
            style_header(c)
        ws6.column_dimensions["A"].width = 8
        ws6.column_dimensions["B"].width = 10
        ws6.column_dimensions["C"].width = 10
        ws6.column_dimensions["D"].width = 80

        transcript_data = analysis_data.get("transcript", {})
        for i, seg in enumerate(transcript_data.get("segments", [])[:500], 2):
            ws6.cell(row=i, column=1, value=seg.get("id", i-1))
            ws6.cell(row=i, column=2, value=f"{seg.get('start', 0):.2f}s")
            ws6.cell(row=i, column=3, value=f"{seg.get('end', 0):.2f}s")
            ws6.cell(row=i, column=4, value=seg.get("text", ""))

        # ── Sheet 7: Generated Script ──────────────────────────────────────
        ws7 = wb.create_sheet("Generated Script")
        ws7.column_dimensions["A"].width = 15
        ws7.column_dimensions["B"].width = 80
        script = analysis_data.get("generated_script") or {}
        row7 = 1
        for key, val in [
            ("Title", script.get("title", "")),
            ("Platform", script.get("platform", "")),
            ("Tone", script.get("tone", "")),
            ("Est. Duration", f"{script.get('estimated_duration', 0)}s"),
        ]:
            ws7.cell(row=row7, column=1, value=key).font = Font(bold=True)
            ws7.cell(row=row7, column=2, value=val)
            row7 += 1

        row7 += 1
        ws7.cell(row=row7, column=1, value="Full Script").font = Font(bold=True)
        row7 += 1
        ws7.merge_cells(start_row=row7, start_column=1, end_row=row7+30, end_column=2)
        full_cell = ws7.cell(row=row7, column=1, value=script.get("full_script", ""))
        full_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Save to bytes
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except ImportError:
        logger.error("openpyxl not installed — cannot generate XLSX.")
        raise
    except Exception as e:
        logger.error(f"XLSX generation failed: {e}")
        raise
